"""Excel parser for the UNAC Maestria wizard flow."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl

from app.core.services.maestria_payload_mapper import map_maestria_values, normalize_maestria_details
from app.core.utils.excel_template_builder import (
    MATRIX_GENERAL_ROW,
    MATRIX_SHEET_NAME,
    MATRIX_SPECIFIC_END_ROW,
    MATRIX_SPECIFIC_START_ROW,
    OPER_VD_ROW_END,
    OPER_VD_ROW_START,
    OPER_VD_SHEET_NAME,
    OPER_VI_ROW_END,
    OPER_VI_ROW_START,
    OPER_VI_SHEET_NAME,
    SHEET_NAME,
)

_KEY_COLUMN = "D"
_VALUE_COLUMN = "B"
_YEAR_MIN = 2000
_YEAR_MAX = 2100
_ORCID_PATTERN = re.compile(r"^\d{4}-\d{4}-\d{4}-\d{3}[\dX]$")
_DNI_PATTERN = re.compile(r"^\d{8}$")
_VALID_ENFOQUE = {"cuantitativo", "cualitativo", "mixto"}
_REQUIRED_FIELDS = {
    "titulo",
    "linea_investigacion",
    "anio",
    "autor1_nombres",
    "asesor_nombres",
    "objeto_estudio",
    "variable_independiente",
    "variable_dependiente",
    "lugar_ejecucion",
    "unidad_analisis",
    "tipo",
    "enfoque",
    "diseno_investigacion",
    "tema_ocde_1",
    "poblacion",
    "muestra",
    "temporal",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"[ \t]+", " ", text).strip()


def _split_lines(value: Any) -> list[str]:
    return [line.strip() for line in _clean(value).split("\n") if line.strip()]


def _extract_labeled_pair(value: Any) -> tuple[str, str]:
    first = ""
    second = ""
    for line in _split_lines(value):
        lower = line.lower()
        if lower.startswith("tecnica:") or lower.startswith("técnica:"):
            first = line.split(":", 1)[1].strip()
        elif lower.startswith("metodo:") or lower.startswith("método:"):
            first = line.split(":", 1)[1].strip()
        elif lower.startswith("metodo y tecnica:") or lower.startswith("método y técnica:"):
            first = line.split(":", 1)[1].strip()
        elif lower.startswith("instrumento:") or lower.startswith("instrumentos:"):
            second = line.split(":", 1)[1].strip()
        elif not first:
            first = line
        elif not second:
            second = line
        else:
            second = f"{second}\n{line}"
    return first, second


@dataclass
class AutorData:
    nombres: str = ""
    dni: str = ""
    orcid: str = ""


@dataclass
class InvestigacionData:
    lugar_ejecucion: str = ""
    unidad_analisis: str = ""
    tipo: str = ""
    enfoque: str = ""
    diseno_investigacion: str = ""
    nivel_investigacion: str = ""
    tema_ocde: list[str] = field(default_factory=list)
    vi: str = ""
    vd: str = ""
    variable_independiente: str = ""
    variable_dependiente: str = ""
    objeto_estudio: str = ""
    poblacion: str = ""
    muestra: str = ""
    lugar: str = ""
    temporal: str = ""


@dataclass
class MaestriaExcelResult:
    titulo: str | None = None
    linea_investigacion: str | None = None
    anio: str | None = None
    lugar_caratula: str | None = None
    autor1: AutorData = field(default_factory=AutorData)
    autor2: AutorData = field(default_factory=AutorData)
    asesor: AutorData = field(default_factory=AutorData)
    coasesor: AutorData = field(default_factory=AutorData)
    investigacion: InvestigacionData = field(default_factory=InvestigacionData)
    matriz_consistencia: dict[str, Any] = field(default_factory=dict)
    operacionalizacion_vi: dict[str, Any] = field(default_factory=dict)
    operacionalizacion_vd: dict[str, Any] = field(default_factory=dict)
    extracted_fields: list[str] = field(default_factory=list)
    missing_required: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    validation_errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "titulo": self.titulo or "",
            "linea_investigacion": self.linea_investigacion or "",
            "anio": self.anio or "",
            "lugar_caratula": self.lugar_caratula or "",
            "autor1_nombres": self.autor1.nombres,
            "autor1_dni": self.autor1.dni,
            "autor1_orcid": self.autor1.orcid,
            "autor2_nombres": self.autor2.nombres,
            "autor2_dni": self.autor2.dni,
            "autor2_orcid": self.autor2.orcid,
            "asesor_nombres": self.asesor.nombres,
            "asesor_dni": self.asesor.dni,
            "asesor_orcid": self.asesor.orcid,
            "coasesor_nombres": self.coasesor.nombres,
            "coasesor_dni": self.coasesor.dni,
            "coasesor_orcid": self.coasesor.orcid,
            "lugar_ejecucion": self.investigacion.lugar_ejecucion,
            "unidad_analisis": self.investigacion.unidad_analisis,
            "tipo": self.investigacion.tipo,
            "enfoque": self.investigacion.enfoque,
            "diseno_investigacion": self.investigacion.diseno_investigacion,
            "nivel_investigacion": self.investigacion.nivel_investigacion,
            "objeto_estudio": self.investigacion.objeto_estudio,
            "variable_independiente": self.investigacion.variable_independiente,
            "variable_dependiente": self.investigacion.variable_dependiente,
            "poblacion": self.investigacion.poblacion,
            "muestra": self.investigacion.muestra,
            "lugar": self.investigacion.lugar,
            "temporal": self.investigacion.temporal,
            "tema_ocde": list(self.investigacion.tema_ocde),
            "tema_ocde_1": self.investigacion.tema_ocde[0] if len(self.investigacion.tema_ocde) > 0 else "",
            "tema_ocde_2": self.investigacion.tema_ocde[1] if len(self.investigacion.tema_ocde) > 1 else "",
            "tema_ocde_3": self.investigacion.tema_ocde[2] if len(self.investigacion.tema_ocde) > 2 else "",
            "matriz_consistencia": self.matriz_consistencia,
            "operacionalizacion_vi": self.operacionalizacion_vi,
            "operacionalizacion_vd": self.operacionalizacion_vd,
        }

    def to_flat_dict(self) -> dict[str, Any]:
        return map_maestria_values(self.to_dict())


def parse_excel_bytes(data: bytes) -> MaestriaExcelResult:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}") from exc

    if SHEET_NAME not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"No se encontró la hoja '{SHEET_NAME}' en el archivo. "
            f"Hojas disponibles: {available}. Usa la plantilla oficial."
        )

    raw_values = _extract_keyed_values(workbook[SHEET_NAME])
    raw_values["matriz_consistencia"] = (
        _parse_matrix_sheet(workbook) if MATRIX_SHEET_NAME in workbook.sheetnames else {}
    )
    raw_values["operacionalizacion_vi"] = (
        _parse_oper_sheet(workbook[OPER_VI_SHEET_NAME], is_vd=False)
        if OPER_VI_SHEET_NAME in workbook.sheetnames
        else {}
    )
    raw_values["operacionalizacion_vd"] = (
        _parse_oper_sheet(workbook[OPER_VD_SHEET_NAME], is_vd=True) if OPER_VD_SHEET_NAME in workbook.sheetnames else {}
    )

    details = normalize_maestria_details(raw_values)
    details["matriz_consistencia"]["dimensiones_variable_independiente"] = _dimensions_from_rows(
        details["matriz_consistencia"].get("dimensiones_variable_independiente"),
        details["operacionalizacion_vi"].get("filas"),
    )
    details["matriz_consistencia"]["dimensiones_variable_dependiente"] = _dimensions_from_rows(
        details["matriz_consistencia"].get("dimensiones_variable_dependiente"),
        details["operacionalizacion_vd"].get("filas"),
    )

    result = _result_from_details(details)
    result.extracted_fields = sorted(
        {
            *(key for key, value in raw_values.items() if value and not isinstance(value, dict)),
            *(_table_fields_present(details)),
        }
    )
    _run_validations(result, raw_values)
    return result


def _extract_keyed_values(sheet: Any) -> dict[str, str]:
    raw: dict[str, str] = {}
    for row in sheet.iter_rows():
        key_cell = None
        value_cell = None
        for cell in row:
            if not hasattr(cell, "column_letter"):
                continue
            if cell.column_letter == _KEY_COLUMN:
                key_cell = cell
            elif cell.column_letter == _VALUE_COLUMN:
                value_cell = cell
        if not key_cell or not value_cell:
            continue
        key = _clean(key_cell.value)
        if key:
            raw[key] = _clean(value_cell.value)
    return raw


def _parse_matrix_sheet(workbook: Any) -> dict[str, Any]:
    sheet = workbook[MATRIX_SHEET_NAME]
    return {
        "problema_general": _clean(sheet[f"A{MATRIX_GENERAL_ROW}"].value),
        "objetivo_general": _clean(sheet[f"B{MATRIX_GENERAL_ROW}"].value),
        "hipotesis_general": _clean(sheet[f"C{MATRIX_GENERAL_ROW}"].value),
        "problemas_especificos": [
            _clean(sheet[f"A{row}"].value)
            for row in range(MATRIX_SPECIFIC_START_ROW, MATRIX_SPECIFIC_END_ROW + 1)
            if _clean(sheet[f"A{row}"].value)
        ],
        "objetivos_especificos": [
            _clean(sheet[f"B{row}"].value)
            for row in range(MATRIX_SPECIFIC_START_ROW, MATRIX_SPECIFIC_END_ROW + 1)
            if _clean(sheet[f"B{row}"].value)
        ],
        "hipotesis_especificas": [
            _clean(sheet[f"C{row}"].value)
            for row in range(MATRIX_SPECIFIC_START_ROW, MATRIX_SPECIFIC_END_ROW + 1)
            if _clean(sheet[f"C{row}"].value)
        ],
        "tecnicas": _clean(sheet["B14"].value),
        "instrumentos": _clean(sheet["B15"].value),
        "procesamiento_datos": _clean(sheet["B16"].value),
    }


def _parse_oper_sheet(sheet: Any, *, is_vd: bool) -> dict[str, Any]:
    row_start = OPER_VD_ROW_START if is_vd else OPER_VI_ROW_START
    row_end = OPER_VD_ROW_END if is_vd else OPER_VI_ROW_END
    rows: list[dict[str, str]] = []
    for row in range(row_start, row_end + 1):
        method_text, instrument_text = _extract_labeled_pair(sheet[f"G{row}"].value)
        item = {
            "dimension": _clean(sheet[f"D{row}"].value),
            "indicador": _clean(sheet[f"E{row}"].value),
            "indice": _clean(sheet[f"F{row}"].value),
            "metodo_tecnica": method_text,
            "tecnica_instrumentos": instrument_text,
        }
        if any(item.values()):
            rows.append(item)
    return {
        "definicion_conceptual": _clean(sheet[f"B{row_start}"].value),
        "definicion_operacional": _clean(sheet[f"C{row_start}"].value),
        "filas": rows,
    }


def _dimensions_from_rows(existing: Any, rows: Any) -> list[str]:
    dimensions = [item.strip() for item in (existing or []) if str(item).strip()]
    if dimensions:
        return dimensions
    return [
        str(item.get("dimension") or "").strip() for item in (rows or []) if str(item.get("dimension") or "").strip()
    ]


def _result_from_details(details: dict[str, Any]) -> MaestriaExcelResult:
    temas = [details.get("tema_ocde_1"), details.get("tema_ocde_2"), details.get("tema_ocde_3")]
    return MaestriaExcelResult(
        titulo=details.get("titulo") or None,
        linea_investigacion=details.get("linea_investigacion") or None,
        anio=details.get("anio") or None,
        lugar_caratula=details.get("lugar_caratula") or None,
        autor1=AutorData(
            details.get("autor1_nombres", ""),
            details.get("autor1_dni", ""),
            details.get("autor1_orcid", ""),
        ),
        autor2=AutorData(
            details.get("autor2_nombres", ""),
            details.get("autor2_dni", ""),
            details.get("autor2_orcid", ""),
        ),
        asesor=AutorData(
            details.get("asesor_nombres", ""),
            details.get("asesor_dni", ""),
            details.get("asesor_orcid", ""),
        ),
        coasesor=AutorData(
            details.get("coasesor_nombres", ""),
            details.get("coasesor_dni", ""),
            details.get("coasesor_orcid", ""),
        ),
        investigacion=InvestigacionData(
            lugar_ejecucion=details.get("lugar_ejecucion", ""),
            unidad_analisis=details.get("unidad_analisis", ""),
            tipo=details.get("tipo", ""),
            enfoque=details.get("enfoque", ""),
            diseno_investigacion=details.get("diseno_investigacion", ""),
            nivel_investigacion=details.get("nivel_investigacion", ""),
            tema_ocde=[item for item in temas if item],
            vi=details.get("variable_independiente", ""),
            vd=details.get("variable_dependiente", ""),
            variable_independiente=details.get("variable_independiente", ""),
            variable_dependiente=details.get("variable_dependiente", ""),
            objeto_estudio=details.get("objeto_estudio", ""),
            poblacion=details.get("poblacion", ""),
            muestra=details.get("muestra", ""),
            lugar=details.get("lugar", ""),
            temporal=details.get("temporal", ""),
        ),
        matriz_consistencia=details.get("matriz_consistencia") or {},
        operacionalizacion_vi=details.get("operacionalizacion_vi") or {},
        operacionalizacion_vd=details.get("operacionalizacion_vd") or {},
    )


def _table_fields_present(details: dict[str, Any]) -> set[str]:
    present: set[str] = set()
    matrix = details.get("matriz_consistencia") or {}
    if any(str(matrix.get(key) or "").strip() for key in ("problema_general", "objetivo_general", "hipotesis_general")):
        present.add("matriz_consistencia")
    if any((matrix.get("problemas_especificos") or [])) or any((matrix.get("objetivos_especificos") or [])):
        present.add("matriz_consistencia")
    if (details.get("operacionalizacion_vi") or {}).get("filas"):
        present.add("operacionalizacion_vi")
    if (details.get("operacionalizacion_vd") or {}).get("filas"):
        present.add("operacionalizacion_vd")
    return present


def _run_validations(result: MaestriaExcelResult, raw_values: dict[str, Any] | None = None) -> None:
    details = result.to_dict()
    missing: list[str] = []
    warnings: list[str] = []
    errors: list[str] = []
    raw_values = raw_values or {}

    for field_key in sorted(_REQUIRED_FIELDS):
        if field_key == "tema_ocde_1":
            if not _clean(raw_values.get("tema_ocde_1")):
                missing.append(field_key)
            continue
        if not _clean(details.get(field_key)):
            missing.append(field_key)

    year = _clean(result.anio)
    if year:
        if not re.fullmatch(r"\d{4}", year):
            errors.append("El año debe contener exactamente 4 dígitos.")
        else:
            year_number = int(year)
            if year_number < _YEAR_MIN or year_number > _YEAR_MAX:
                errors.append(f"El año debe estar dentro del rango {_YEAR_MIN}-{_YEAR_MAX}.")

    for label, value in (
        ("Autor 1", result.autor1.orcid),
        ("Autor 2", result.autor2.orcid),
        ("Asesor", result.asesor.orcid),
    ):
        if value and not _ORCID_PATTERN.match(value):
            warnings.append(f"ORCID de {label} '{value}' no tiene el formato esperado (0000-0000-0000-0000).")

    for label, value in (
        ("Autor 1", result.autor1.dni),
        ("Autor 2", result.autor2.dni),
        ("Asesor", result.asesor.dni),
    ):
        if value and not _DNI_PATTERN.match(value):
            warnings.append(f"DNI de {label} '{value}' no parece un DNI peruano válido (8 dígitos).")

    if result.investigacion.enfoque and result.investigacion.enfoque.lower() not in _VALID_ENFOQUE:
        warnings.append(
            f"Enfoque '{result.investigacion.enfoque}' no es uno de los valores esperados: "
            f"{', '.join(sorted(_VALID_ENFOQUE))}."
        )

    autor2_values = [result.autor2.nombres, result.autor2.dni, result.autor2.orcid]
    if 0 < len([item for item in autor2_values if item]) < 2:
        warnings.append("Autor 2: si va a incluir un segundo autor, se recomienda completar al menos nombres y DNI.")

    result.missing_required = missing
    result.warnings = warnings
    result.validation_errors = errors
