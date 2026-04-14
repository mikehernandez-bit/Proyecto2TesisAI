"""Excel template builder for the UNAC Maestria wizard flow."""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SHEET_NAME = "Datos Maestría"
GUIDE_SHEET_NAME = "Guía"
MATRIX_SHEET_NAME = "Matriz de consistencia"
OPER_VI_SHEET_NAME = "Operacionalización VI"
OPER_VD_SHEET_NAME = "Operacionalización VD"

_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_SOFT_FILL = PatternFill("solid", fgColor="F8FAFC")
_EDITABLE_FILL = PatternFill("solid", fgColor="FEF3C7")
_AUTO_FILL = PatternFill("solid", fgColor="E2E8F0")
_TABLE_HEADER_FILL = PatternFill("solid", fgColor="334155")
_SECTION_FILL = PatternFill("solid", fgColor="2563EB")
_SECTION_ALT_FILL = PatternFill("solid", fgColor="0F766E")
_SECTION_WARN_FILL = PatternFill("solid", fgColor="7C3AED")

_SCALAR_SECTIONS: list[dict[str, Any]] = [
    {
        "title": "A. Datos generales",
        "fill": _SECTION_FILL,
        "fields": [
            (
                "titulo",
                "Título del proyecto *",
                "Ej: Implementación de un sistema para mejorar la atención de proyectos.",
            ),
            (
                "linea_investigacion",
                "Línea de investigación *",
                "Ej: Gerencia y optimización de sistemas de mantenimiento.",
            ),
            ("anio", "Año *", "Ej: 2026"),
            ("lugar_caratula", "Lugar de carátula", "Ej: Callao"),
        ],
    },
    {
        "title": "B. Autor 1",
        "fill": _SECTION_ALT_FILL,
        "fields": [
            (
                "autor1_nombres",
                "Apellidos y nombres (Autor 1) *",
                "Ej: QUISPE FLORES, Juan Carlos",
            ),
            ("autor1_dni", "DNI (Autor 1)", "Ej: 12345678"),
            ("autor1_orcid", "ORCID (Autor 1)", "Ej: 0000-0001-2345-6789"),
        ],
    },
    {
        "title": "C. Autor 2 (opcional)",
        "fill": _SECTION_ALT_FILL,
        "fields": [
            (
                "autor2_nombres",
                "Apellidos y nombres (Autor 2)",
                "Déjalo vacío si no hay segundo autor.",
            ),
            ("autor2_dni", "DNI (Autor 2)", "Ej: 87654321"),
            ("autor2_orcid", "ORCID (Autor 2)", "Ej: 0000-0002-3456-7890"),
        ],
    },
    {
        "title": "D. Asesor",
        "fill": _SECTION_FILL,
        "fields": [
            (
                "asesor_nombres",
                "Apellidos y nombres (Asesor) *",
                "Ej: RAMÍREZ TORRES, Pedro Augusto",
            ),
            ("asesor_dni", "DNI (Asesor)", "Ej: 11223344"),
            ("asesor_orcid", "ORCID (Asesor)", "Ej: 0000-0003-4567-8901"),
        ],
    },
    {
        "title": "E. Datos de investigación",
        "fill": _SECTION_WARN_FILL,
        "fields": [
            ("objeto_estudio", "Objeto de estudio *", "Ej: Flota de motoniveladoras CAT 24M."),
            (
                "variable_independiente",
                "Variable independiente *",
                "Ej: Plan de mantenimiento centrado en confiabilidad.",
            ),
            ("variable_dependiente", "Variable dependiente *", "Ej: Disponibilidad inherente."),
            (
                "lugar_ejecucion",
                "Lugar de ejecución *",
                "Ej: Unidad minera cuprífera, Sierra Central.",
            ),
            ("unidad_analisis", "Unidad de análisis *", "Ej: Motoniveladoras CAT 24M."),
            ("tipo", "Tipo de investigación *", "Ej: Aplicada"),
            ("enfoque", "Enfoque *", "Ej: Cuantitativo"),
            ("diseno_investigacion", "Diseño de investigación *", "Ej: Pre experimental"),
            ("nivel_investigacion", "Nivel de investigación", "Ej: Explicativa"),
            ("tema_ocde_1", "Tema OCDE 1 *", "Ej: 2.00.00 -- Ingeniería, Tecnología"),
            ("tema_ocde_2", "Tema OCDE 2", "Ej: 2.03.00 -- Ingeniería Mecánica"),
            ("tema_ocde_3", "Tema OCDE 3", "Ej: 2.03.06 -- Análisis de confiabilidad"),
            ("poblacion", "Población *", "Ej: 05 motoniveladoras Caterpillar (CAT) modelo 24M"),
            ("muestra", "Muestra *", "Ej: Muestreo no probabilístico tipo censal"),
            ("lugar", "Lugar (para el título)", "Ej: Sierra Central"),
            ("temporal", "Temporal *", "Ej: 2025"),
        ],
    },
]

MATRIX_GENERAL_ROW = 6
MATRIX_SPECIFIC_START_ROW = 8
MATRIX_SPECIFIC_END_ROW = 11
OPER_VI_ROW_START = 4
OPER_VI_ROW_END = 7
OPER_VD_ROW_START = 4
OPER_VD_ROW_END = 5


def build_excel_template() -> bytes:
    """Build and return the Maestria Excel template as XLSX bytes."""
    workbook = openpyxl.Workbook()

    guide_sheet = workbook.active
    guide_sheet.title = GUIDE_SHEET_NAME

    data_sheet = workbook.create_sheet(SHEET_NAME)
    field_rows = _write_data_sheet(data_sheet)
    _write_guide_sheet(guide_sheet)

    matrix_sheet = workbook.create_sheet(MATRIX_SHEET_NAME)
    _write_matrix_sheet(matrix_sheet, field_rows)

    oper_vi_sheet = workbook.create_sheet(OPER_VI_SHEET_NAME)
    _write_operationalization_sheet(oper_vi_sheet, field_rows, is_vd=False)

    oper_vd_sheet = workbook.create_sheet(OPER_VD_SHEET_NAME)
    _write_operationalization_sheet(oper_vd_sheet, field_rows, is_vd=True)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _write_data_sheet(sheet: Any) -> dict[str, int]:
    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 50
    sheet.column_dimensions["C"].width = 54
    sheet.column_dimensions["D"].hidden = True

    sheet.merge_cells("A1:C1")
    _set_cell(
        sheet["A1"],
        "PLANTILLA TESIS DE MAESTRÍA UNAC",
        font=Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        fill=_HEADER_FILL,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:C2")
    _set_cell(
        sheet["A2"],
        (
            "Completa primero los datos base. Las hojas de Matriz y Operacionalización "
            "reutilizan automáticamente el título, las variables y la metodología base."
        ),
        font=Font(name="Arial", size=10, italic=True, color="374151"),
        fill=PatternFill("solid", fgColor="EEF6FF"),
        alignment=Alignment(wrap_text=True, vertical="center"),
        border=_THIN_BORDER,
    )
    sheet.row_dimensions[2].height = 34

    _table_header(sheet, 3, ["Campo", "Valor", "Ejemplo / nota"])

    field_rows: dict[str, int] = {}
    row = 4
    for section in _SCALAR_SECTIONS:
        sheet.merge_cells(f"A{row}:C{row}")
        _set_cell(
            sheet[f"A{row}"],
            section["title"],
            font=Font(name="Arial", size=11, bold=True, color="FFFFFF"),
            fill=section["fill"],
            alignment=Alignment(vertical="center"),
            border=_THIN_BORDER,
        )
        sheet.row_dimensions[row].height = 24
        row += 1

        for key, label, example in section["fields"]:
            field_rows[key] = row
            _set_cell(
                sheet[f"A{row}"],
                label,
                font=Font(name="Arial", size=10, bold=label.endswith("*")),
                fill=_SOFT_FILL,
                alignment=Alignment(wrap_text=True, vertical="center"),
                border=_THIN_BORDER,
            )
            _set_cell(
                sheet[f"B{row}"],
                "",
                font=Font(name="Arial", size=10),
                fill=PatternFill("solid", fgColor="FFFFFF"),
                alignment=Alignment(wrap_text=True, vertical="top"),
                border=_THIN_BORDER,
            )
            _set_cell(
                sheet[f"C{row}"],
                example,
                font=Font(name="Arial", size=9, italic=True, color="6B7280"),
                fill=_SOFT_FILL,
                alignment=Alignment(wrap_text=True, vertical="top"),
                border=_THIN_BORDER,
            )
            _set_cell(sheet[f"D{row}"], key)
            sheet.row_dimensions[row].height = 24
            row += 1

        row += 1

    return field_rows


def _write_guide_sheet(sheet: Any) -> None:
    sheet.column_dimensions["A"].width = 98
    lines = [
        ("CÓMO USAR ESTA PLANTILLA", True, "1E3A5F"),
        ("", False, ""),
        ("1. Completa primero la hoja 'Datos Maestría'.", False, ""),
        ("2. Luego llena 'Matriz de consistencia', 'Operacionalización VI' y 'Operacionalización VD'.", False, ""),
        (
            "3. Amarillo = lo llena el usuario. Gris = se completa automáticamente desde Datos Maestría.",
            False,
            "",
        ),
        ("4. No se incluye índice de abreviaturas en esta plantilla.", False, ""),
        (
            "5. Operacionalización VI tiene 4 filas de dimensiones; Operacionalización VD tiene 2.",
            False,
            "",
        ),
        (
            "6. Si una tabla queda incompleta, podrás corregirla manualmente en el Paso 3 del wizard.",
            False,
            "",
        ),
        ("", False, ""),
        ("FORMATO DE CAPTURA", True, "2563EB"),
        (
            "- Matriz de consistencia: llena solo las celdas amarillas de problema, objetivo e hipótesis.",
            False,
            "",
        ),
        (
            "- Variables y metodología se reflejan automáticamente; "
            "solo técnicas, instrumentos y procesamiento se editan a mano.",
            False,
            "",
        ),
        ("- Operacionalización VI: bloque fijo de variable + 4 filas por dimensión.", False, ""),
        ("- Operacionalización VD: bloque fijo de variable + 2 filas por dimensión.", False, ""),
    ]

    for index, (text, is_header, color) in enumerate(lines, start=1):
        cell = sheet[f"A{index}"]
        if is_header:
            _set_cell(
                cell,
                text,
                font=Font(name="Arial", size=12 if index == 1 else 11, bold=True, color="FFFFFF"),
                fill=PatternFill("solid", fgColor=color),
                alignment=Alignment(vertical="center"),
                border=_THIN_BORDER,
            )
            sheet.row_dimensions[index].height = 24
        else:
            _set_cell(
                cell,
                text,
                font=Font(name="Arial", size=10),
                alignment=Alignment(wrap_text=True, vertical="top"),
            )


def _write_matrix_sheet(sheet: Any, field_rows: dict[str, int]) -> None:
    sheet.freeze_panes = "A5"
    for column, width in {"A": 34, "B": 34, "C": 34, "D": 28, "E": 30}.items():
        sheet.column_dimensions[column].width = width

    sheet.merge_cells("A1:E1")
    _set_cell(
        sheet["A1"],
        "Anexo 1: Matriz de consistencia",
        font=Font(name="Arial", size=13, bold=True, color="FFFFFF"),
        fill=_SECTION_WARN_FILL,
        alignment=Alignment(horizontal="center", vertical="center"),
        border=_THIN_BORDER,
    )
    sheet.row_dimensions[1].height = 26

    sheet.merge_cells("A2:E2")
    _set_cell(
        sheet["A2"],
        _ref_formula(SHEET_NAME, field_rows["titulo"], default=""),
        font=Font(name="Arial", size=11, bold=True, color="111827"),
        fill=PatternFill("solid", fgColor="F8FAFC"),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=_THIN_BORDER,
    )
    sheet.row_dimensions[2].height = 32

    sheet.merge_cells("A3:E3")
    _set_cell(
        sheet["A3"],
        (
            "Completa solo las celdas amarillas. Las celdas grises se reflejan automáticamente "
            "desde Datos Maestría y las hojas de operacionalización."
        ),
        font=Font(name="Arial", size=10, italic=True, color="374151"),
        fill=PatternFill("solid", fgColor="EEF6FF"),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=_THIN_BORDER,
    )
    sheet.row_dimensions[3].height = 36

    _table_header(sheet, 4, ["PROBLEMA", "OBJETIVOS", "HIPÓTESIS", "VARIABLES", "METODOLOGÍA"])

    _matrix_subheader(sheet, "A5", "PROBLEMA GENERAL")
    _matrix_subheader(sheet, "B5", "OBJETIVO GENERAL")
    _matrix_subheader(sheet, "C5", "HIPÓTESIS GENERAL")
    sheet.merge_cells("D5:D11")
    sheet.merge_cells("E5:E11")
    _formula_cell(
        sheet,
        "D5",
        _build_matrix_variables_formula(field_rows),
        alignment=Alignment(horizontal="center", vertical="top", wrap_text=True),
    )
    _formula_cell(
        sheet,
        "E5",
        _build_matrix_methodology_formula(field_rows),
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )
    sheet.row_dimensions[5].height = 80

    _editable_cell(sheet, "A6", comment="Escribe aquí el problema general.")
    _editable_cell(sheet, "B6", comment="Escribe aquí el objetivo general.")
    _editable_cell(sheet, "C6", comment="Escribe aquí la hipótesis general.")

    _matrix_subheader(sheet, "A7", "PROBLEMAS ESPECÍFICOS")
    _matrix_subheader(sheet, "B7", "OBJETIVOS ESPECÍFICOS")
    _matrix_subheader(sheet, "C7", "HIPÓTESIS ESPECÍFICAS")

    for row in range(MATRIX_SPECIFIC_START_ROW, MATRIX_SPECIFIC_END_ROW + 1):
        number = row - MATRIX_SPECIFIC_START_ROW + 1
        _editable_cell(
            sheet,
            f"A{row}",
            comment=f"Escribe aquí el problema específico {number}.",
        )
        _editable_cell(
            sheet,
            f"B{row}",
            comment=f"Escribe aquí el objetivo específico {number}.",
        )
        _editable_cell(
            sheet,
            f"C{row}",
            comment=f"Escribe aquí la hipótesis específica {number}.",
        )

    sheet.merge_cells("A13:E13")
    _set_cell(
        sheet["A13"],
        "Campos de apoyo para Metodología. Lo que escribas aquí se refleja automáticamente en la columna 5.",
        font=Font(name="Arial", size=10, italic=True, color="374151"),
        fill=PatternFill("solid", fgColor="EEF6FF"),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=_THIN_BORDER,
    )
    sheet.row_dimensions[13].height = 30

    _matrix_subheader(sheet, "A14", "TÉCNICAS")
    sheet.merge_cells("B14:E14")
    _editable_cell(sheet, "B14", comment="Escribe aquí las técnicas de investigación que utilizarás.")
    _matrix_subheader(sheet, "A15", "INSTRUMENTOS")
    sheet.merge_cells("B15:E15")
    _editable_cell(sheet, "B15", comment="Escribe aquí los instrumentos que usarás para recolectar datos.")
    _matrix_subheader(sheet, "A16", "PROCESAMIENTO DE DATOS")
    sheet.merge_cells("B16:E16")
    _editable_cell(sheet, "B16", comment="Escribe aquí cómo procesarás y analizarás los datos.")


def _write_operationalization_sheet(sheet: Any, field_rows: dict[str, int], *, is_vd: bool) -> None:
    title = "Operacionalización de variable dependiente" if is_vd else "Operacionalización de variable independiente"
    sheet.freeze_panes = "A3"
    for column, width in {
        "A": 24,
        "B": 34,
        "C": 34,
        "D": 24,
        "E": 24,
        "F": 18,
        "G": 34,
    }.items():
        sheet.column_dimensions[column].width = width

    sheet.merge_cells("A1:G1")
    _set_cell(
        sheet["A1"],
        title,
        font=Font(name="Arial", size=13, bold=True, color="FFFFFF"),
        fill=_SECTION_ALT_FILL if is_vd else _SECTION_FILL,
        alignment=Alignment(horizontal="center", vertical="center"),
        border=_THIN_BORDER,
    )
    sheet.row_dimensions[1].height = 26

    note = (
        "Completa solo las filas amarillas de dimensiones. "
        + ("Esta hoja tiene 2 filas de captura." if is_vd else "Esta hoja tiene 4 filas de captura.")
        + " La variable y definiciones se llenan una sola vez en el bloque izquierdo."
    )
    sheet.merge_cells("A2:G2")
    _set_cell(
        sheet["A2"],
        note,
        font=Font(name="Arial", size=10, italic=True, color="374151"),
        fill=PatternFill("solid", fgColor="EEF6FF"),
        alignment=Alignment(wrap_text=True, vertical="center"),
        border=_THIN_BORDER,
    )
    sheet.row_dimensions[2].height = 34

    last_header = "MÉTODO Y TÉCNICA" if is_vd else "TÉCNICA E INSTRUMENTOS"
    _table_header(
        sheet,
        3,
        [
            "VARIABLE",
            "DEFINICIÓN CONCEPTUAL",
            "DEFINICIÓN OPERACIONAL",
            "DIMENSIONES",
            "INDICADORES",
            "ÍNDICE",
            last_header,
        ],
    )

    variable_key = "variable_dependiente" if is_vd else "variable_independiente"
    row_start = OPER_VD_ROW_START if is_vd else OPER_VI_ROW_START
    row_end = OPER_VD_ROW_END if is_vd else OPER_VI_ROW_END
    sheet.merge_cells(f"A{row_start}:A{row_end}")
    sheet.merge_cells(f"B{row_start}:B{row_end}")
    sheet.merge_cells(f"C{row_start}:C{row_end}")

    _formula_cell(
        sheet,
        f"A{row_start}",
        _ref_formula(
            SHEET_NAME,
            field_rows[variable_key],
            default="Completa la variable en Datos Maestría",
        ),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=_AUTO_FILL,
    )
    _editable_cell(
        sheet,
        f"B{row_start}",
        alignment=Alignment(wrap_text=True, vertical="top"),
        comment="Escribe aquí la definición conceptual de la variable.",
    )
    _editable_cell(
        sheet,
        f"C{row_start}",
        alignment=Alignment(wrap_text=True, vertical="top"),
        comment="Escribe aquí la definición operacional de la variable.",
    )

    for row in range(row_start, row_end + 1):
        _editable_cell(sheet, f"D{row}", comment="Escribe aquí una dimensión.")
        _editable_cell(sheet, f"E{row}", comment="Escribe aquí el indicador asociado a la dimensión.")
        _editable_cell(sheet, f"F{row}", comment="Escribe aquí el índice o escala de medición.")
        _editable_cell(
            sheet,
            f"G{row}",
            alignment=Alignment(wrap_text=True, vertical="top"),
            comment=(
                "Escribe aquí la técnica y el instrumento. "
                "Puedes usar dos líneas: 'Técnica: ...' e 'Instrumento: ...'."
            ),
        )
        sheet.row_dimensions[row].height = 52


def _table_header(sheet: Any, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        _set_cell(
            sheet.cell(row=row, column=col),
            header,
            font=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
            fill=_TABLE_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=_THIN_BORDER,
        )
    sheet.row_dimensions[row].height = 24


def _matrix_subheader(sheet: Any, coord: str, value: str) -> None:
    _set_cell(
        sheet[coord],
        value,
        font=Font(name="Arial", size=9, bold=True, color="1F2937"),
        fill=_SOFT_FILL,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=_THIN_BORDER,
    )


def _editable_cell(
    sheet: Any,
    coord: str,
    *,
    alignment: Alignment | None = None,
    comment: str | None = None,
) -> None:
    _set_cell(
        sheet[coord],
        "",
        font=Font(name="Arial", size=10),
        fill=_EDITABLE_FILL,
        alignment=alignment or Alignment(wrap_text=True, vertical="top"),
        border=_THIN_BORDER,
    )
    if comment:
        sheet[coord].comment = Comment(comment, "GicaGen")


def _formula_cell(
    sheet: Any,
    coord: str,
    formula: str,
    *,
    alignment: Alignment | None = None,
    fill: PatternFill | None = None,
) -> None:
    _set_cell(
        sheet[coord],
        formula,
        font=Font(name="Arial", size=10),
        fill=fill or _AUTO_FILL,
        alignment=alignment or Alignment(wrap_text=True, vertical="center"),
        border=_THIN_BORDER,
    )


def _ref_formula(sheet_name: str, row: int, *, column: str = "B", default: str = "") -> str:
    ref = f"'{sheet_name}'!{column}{row}"
    escaped_default = default.replace('"', '""')
    return f'=IF({ref}="","{escaped_default}",{ref})'


def _ref_value(sheet_name: str, row: int, *, column: str = "B") -> str:
    return f"'{sheet_name}'!{column}{row}"


def _quoted(text: str) -> str:
    escaped = text.replace('"', '""')
    return f'"{escaped}"'


def _if_value(ref: str, default: str = "") -> str:
    escaped_default = default.replace('"', '""')
    return f'IF({ref}="","{escaped_default}",{ref})'


def _build_matrix_variables_formula(field_rows: dict[str, int]) -> str:
    variable_independiente = _ref_value(SHEET_NAME, field_rows["variable_independiente"])
    variable_dependiente = _ref_value(SHEET_NAME, field_rows["variable_dependiente"])
    parts = [
        _quoted("VARIABLE INDEPENDIENTE"),
        _if_value(variable_independiente, "Completa la variable independiente en Datos Maestría"),
        _quoted("Dimensiones"),
    ]
    parts.extend(
        _if_value(_ref_value(OPER_VI_SHEET_NAME, row, column="D"))
        for row in range(OPER_VI_ROW_START, OPER_VI_ROW_END + 1)
    )
    parts.extend(
        [
            _quoted("VARIABLE DEPENDIENTE"),
            _if_value(variable_dependiente, "Completa la variable dependiente en Datos Maestría"),
            _quoted("Dimensiones"),
        ]
    )
    parts.extend(
        _if_value(_ref_value(OPER_VD_SHEET_NAME, row, column="D"))
        for row in range(OPER_VD_ROW_START, OPER_VD_ROW_END + 1)
    )
    return "=" + "&CHAR(10)&".join(parts)


def _build_matrix_methodology_formula(field_rows: dict[str, int]) -> str:
    type_ref = _ref_value(SHEET_NAME, field_rows["tipo"])
    level_ref = _ref_value(SHEET_NAME, field_rows["nivel_investigacion"])
    focus_ref = _ref_value(SHEET_NAME, field_rows["enfoque"])
    design_ref = _ref_value(SHEET_NAME, field_rows["diseno_investigacion"])
    population_ref = _ref_value(SHEET_NAME, field_rows["poblacion"])
    sample_ref = _ref_value(SHEET_NAME, field_rows["muestra"])
    techniques_ref = _ref_value(MATRIX_SHEET_NAME, 14, column="B")
    instruments_ref = _ref_value(MATRIX_SHEET_NAME, 15, column="B")
    processing_ref = _ref_value(MATRIX_SHEET_NAME, 16, column="B")
    parts = [
        _quoted("Tipo de investigación:"),
        _if_value(type_ref),
        _quoted("Nivel de investigación:"),
        _if_value(level_ref),
        _quoted("Enfoque de investigación:"),
        _if_value(focus_ref),
        _quoted("Diseño:"),
        _if_value(design_ref),
        _quoted("Población:"),
        _if_value(population_ref),
        _quoted("Muestra:"),
        _if_value(sample_ref),
        _quoted("Técnicas:"),
        _if_value(techniques_ref),
        _quoted("Instrumentos:"),
        _if_value(instruments_ref),
        _quoted("Procesamiento de datos:"),
        _if_value(processing_ref),
    ]
    return "=" + "&CHAR(10)&".join(parts)


def _set_cell(
    cell: Any,
    value: Any = None,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
