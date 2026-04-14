"""
Tests for the Maestría UNAC Excel Parser.

Covers:
- Happy path: full Excel template parsed correctly
- Missing required fields
- ORCID validation
- DNI validation
- Year validation
- Enfoque validation
- Autor 2 coherence check
- Invalid file (not xlsx)
- Missing expected sheet
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.core.services.maestria_excel_parser import (
    SHEET_NAME,
    MaestriaExcelResult,
    parse_excel_bytes,
)
from app.core.utils.excel_template_builder import (
    GUIDE_SHEET_NAME,
    MATRIX_SHEET_NAME,
    OPER_VD_SHEET_NAME,
    OPER_VI_SHEET_NAME,
    build_excel_template,
)

# ---------------------------------------------------------------------------
# Helper: build Excel bytes with custom row key→value pairs
# ---------------------------------------------------------------------------


def _build_excel_with_data(data: dict[str, str]) -> bytes:
    """
    Build a minimal Excel file that mimics the template structure.
    Column D = field key, Column B = user value.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for row_idx, (key, value) in enumerate(data.items(), start=3):
        ws[f"D{row_idx}"] = key
        ws[f"B{row_idx}"] = value

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _full_data() -> dict[str, str]:
    return {
        "titulo": "Sistema de Gestión de Mantenimiento",
        "linea_investigacion": "Gestión y Optimización de Sistemas",
        "anio": "2025",
        "lugar_caratula": "Callao",
        "autor1_nombres": "QUISPE FLORES, Juan Carlos",
        "autor1_dni": "12345678",
        "autor1_orcid": "0000-0001-2345-6789",
        "autor2_nombres": "",
        "autor2_dni": "",
        "autor2_orcid": "",
        "asesor_nombres": "RAMÍREZ TORRES, Pedro Augusto",
        "asesor_dni": "11223344",
        "asesor_orcid": "0000-0003-4567-8901",
        "lugar_ejecucion": "Planta Industrial XYZ, Callao",
        "unidad_analisis": "Equipos de mantenimiento",
        "tipo": "Aplicada",
        "enfoque": "Cuantitativo",
        "diseno_investigacion": "Preexperimental",
        "tema_ocde_1": "2. Ingeniería y Tecnología",
        "tema_ocde_2": "2.3 Ingeniería Mecánica",
        "tema_ocde_3": "",
        "objeto_estudio": "Flota de motoniveladoras CAT 24M",
        "variable_independiente": "Plan de mantenimiento centrado en confiabilidad",
        "variable_dependiente": "Disponibilidad inherente",
        "poblacion": "5 motoniveladoras CAT 24M",
        "muestra": "Muestreo no probabilistico de tipo censal",
        "temporal": "2025",
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestParseExcelBytesHappyPath:
    def test_parses_titulo(self) -> None:
        data = _full_data()
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert result.titulo == "Sistema de Gestión de Mantenimiento"

    def test_parses_anio(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert result.anio == "2025"

    def test_parses_autor1_nombres(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert result.autor1.nombres == "QUISPE FLORES, Juan Carlos"

    def test_parses_asesor_nombres(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert result.asesor.nombres == "RAMÍREZ TORRES, Pedro Augusto"

    def test_parses_investigacion_fields(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert result.investigacion.tipo == "Aplicada"
        assert result.investigacion.enfoque == "Cuantitativo"
        assert result.investigacion.diseno_investigacion == "Preexperimental"

    def test_parses_tema_ocde(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert "2. Ingeniería y Tecnología" in result.investigacion.tema_ocde

    def test_no_missing_required_on_full_data(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert len(result.missing_required) == 0

    def test_no_validation_errors_on_full_data(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert len(result.validation_errors) == 0

    def test_extracted_fields_populated(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        assert "titulo" in result.extracted_fields
        assert "autor1_nombres" in result.extracted_fields
        assert "asesor_nombres" in result.extracted_fields

    def test_to_flat_dict_has_title_alias(self) -> None:
        result = parse_excel_bytes(_build_excel_with_data(_full_data()))
        flat = result.to_flat_dict()
        assert flat["title"] == flat["titulo"]
        assert flat["tema"] == flat["titulo"]


class TestParseExcelBytesMissingRequired:
    def test_missing_titulo_is_reported(self) -> None:
        data = _full_data()
        data["titulo"] = ""
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert "titulo" in result.missing_required

    def test_missing_autor1_nombres_is_reported(self) -> None:
        data = _full_data()
        data["autor1_nombres"] = ""
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert "autor1_nombres" in result.missing_required

    def test_missing_anio_is_reported(self) -> None:
        data = _full_data()
        data["anio"] = ""
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert "anio" in result.missing_required

    def test_missing_tema_ocde_1_is_reported(self) -> None:
        data = _full_data()
        data["tema_ocde_1"] = ""
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert "tema_ocde_1" in result.missing_required


class TestParseExcelBytesValidations:
    def test_invalid_year_letters_produces_error(self) -> None:
        data = _full_data()
        data["anio"] = "ABCD"
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert len(result.validation_errors) > 0
        assert any("año" in e.lower() for e in result.validation_errors)

    def test_year_out_of_range_produces_error(self) -> None:
        data = _full_data()
        data["anio"] = "1899"
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert any("rango" in e.lower() for e in result.validation_errors)

    def test_invalid_orcid_format_produces_warning(self) -> None:
        data = _full_data()
        data["autor1_orcid"] = "INVALID-ORCID"
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert any("orcid" in w.lower() and "autor 1" in w.lower() for w in result.warnings)

    def test_valid_orcid_no_warning(self) -> None:
        data = _full_data()
        data["autor1_orcid"] = "0000-0001-2345-6789"
        result = parse_excel_bytes(_build_excel_with_data(data))
        orcid_warnings = [w for w in result.warnings if "orcid" in w.lower() and "autor 1" in w.lower()]
        assert len(orcid_warnings) == 0

    def test_invalid_dni_format_produces_warning(self) -> None:
        data = _full_data()
        data["autor1_dni"] = "1234"  # too short
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert any("dni" in w.lower() and "autor 1" in w.lower() for w in result.warnings)

    def test_valid_dni_no_warning(self) -> None:
        data = _full_data()
        data["autor1_dni"] = "12345678"
        result = parse_excel_bytes(_build_excel_with_data(data))
        dni_warnings = [w for w in result.warnings if "dni" in w.lower() and "autor 1" in w.lower()]
        assert len(dni_warnings) == 0

    def test_unknown_enfoque_produces_warning(self) -> None:
        data = _full_data()
        data["enfoque"] = "Desconocido"
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert any("enfoque" in w.lower() for w in result.warnings)

    def test_known_enfoque_no_warning(self) -> None:
        for enfoque in ("Cuantitativo", "Cualitativo", "Mixto"):
            data = _full_data()
            data["enfoque"] = enfoque
            result = parse_excel_bytes(_build_excel_with_data(data))
            enfoque_warnings = [w for w in result.warnings if "enfoque" in w.lower()]
            assert len(enfoque_warnings) == 0, f"Unexpected warning for enfoque={enfoque}"

    def test_partial_autor2_produces_coherence_warning(self) -> None:
        data = _full_data()
        data["autor2_nombres"] = "PEREZ, Ana"  # only name, no DNI/ORCID
        result = parse_excel_bytes(_build_excel_with_data(data))
        assert any("autor 2" in w.lower() for w in result.warnings)


class TestParseExcelBytesErrorCases:
    def test_not_xlsx_bytes_raises_value_error(self) -> None:
        with pytest.raises(ValueError, match="leer"):
            parse_excel_bytes(b"this is not an xlsx file")

    def test_empty_bytes_raises_value_error(self) -> None:
        with pytest.raises((ValueError, Exception)):
            parse_excel_bytes(b"")

    def test_wrong_sheet_name_raises_value_error(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja incorrecta"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with pytest.raises(ValueError, match="hoja"):
            parse_excel_bytes(buf.read())


class TestOfficialTemplateRoundtrip:
    """Verify that the official template can be parsed without errors."""

    def test_blank_template_parses_without_crash(self) -> None:
        template_bytes = build_excel_template()
        # Blank template has no values — all required fields will be missing
        result = parse_excel_bytes(template_bytes)
        assert isinstance(result, MaestriaExcelResult)

    def test_blank_template_has_missing_required(self) -> None:
        template_bytes = build_excel_template()
        result = parse_excel_bytes(template_bytes)
        # All required fields should be missing since the template is blank
        assert len(result.missing_required) > 0

    def test_blank_template_no_validation_errors(self) -> None:
        """Blank template should not produce validation errors (no values to validate)."""
        template_bytes = build_excel_template()
        result = parse_excel_bytes(template_bytes)
        # Year is blank so no year error
        assert len(result.validation_errors) == 0


class TestOfficialTemplateStructure:
    def test_template_has_expected_sheets_without_abbreviations(self) -> None:
        workbook = openpyxl.load_workbook(io.BytesIO(build_excel_template()))
        assert workbook.sheetnames == [
            GUIDE_SHEET_NAME,
            SHEET_NAME,
            MATRIX_SHEET_NAME,
            OPER_VI_SHEET_NAME,
            OPER_VD_SHEET_NAME,
        ]
        assert all("abreviaturas" not in name.lower() for name in workbook.sheetnames)

    def test_matrix_sheet_explains_editable_vs_automatic_cells(self) -> None:
        workbook = openpyxl.load_workbook(io.BytesIO(build_excel_template()))
        sheet = workbook[MATRIX_SHEET_NAME]
        assert "celdas amarillas" in str(sheet["A3"].value).lower()
        assert "celdas grises" in str(sheet["A3"].value).lower()
        merged_ranges = {str(item) for item in sheet.merged_cells.ranges}
        assert "D5:D11" in merged_ranges
        assert "E5:E11" in merged_ranges
        assert "A2:E2" in merged_ranges

    def test_operacionalizacion_vi_has_four_capture_rows(self) -> None:
        workbook = openpyxl.load_workbook(io.BytesIO(build_excel_template()))
        sheet = workbook[OPER_VI_SHEET_NAME]
        assert sheet["D4"].value is None
        assert sheet["D7"].value is None
        assert sheet["D8"].value is None
        assert "4 filas" in str(sheet["A2"].value)

    def test_operacionalizacion_vd_has_two_capture_rows(self) -> None:
        workbook = openpyxl.load_workbook(io.BytesIO(build_excel_template()))
        sheet = workbook[OPER_VD_SHEET_NAME]
        assert sheet["D4"].value is None
        assert sheet["D5"].value is None
        assert sheet["D6"].value is None
        assert "2 filas" in str(sheet["A2"].value)

    def test_matrix_support_fields_live_below_the_table(self) -> None:
        workbook = openpyxl.load_workbook(io.BytesIO(build_excel_template()))
        sheet = workbook[MATRIX_SHEET_NAME]
        assert sheet["A14"].value == "TÉCNICAS"
        assert sheet["A15"].value == "INSTRUMENTOS"
        assert sheet["A16"].value == "PROCESAMIENTO DE DATOS"
